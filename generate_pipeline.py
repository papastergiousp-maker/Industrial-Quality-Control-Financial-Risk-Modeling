"""
Industrial Tomato Quality Control Pipeline
==========================================
Φορτώνει/γεννά CSV με ωριαία data 8 παραμέτρων ποιότητας,
ταξινομεί κάθε παρτίδα (Accept / Correction / Reject) και
υπολογίζει το κόστος επεξεργασίας + waste κόστος.

Spec sheet: MFD-024 (Νομικός D15 - Σάλτσα Τομάτας / Industrial Tomato Paste)
"""

import os
import math
import csv
from datetime import datetime, timedelta

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage

# -----------------------------------------------------------------------------
# 1. SPEC LIMITS (από τα MFD-024 spec sheets)
# -----------------------------------------------------------------------------
SPEC = {
    "BRIX":         {"min": 28.0, "max": 30.0, "target": 28.5, "unit": "%"},
    "COLOR_a_b":    {"min": 1.85, "max": 2.40, "target": 2.00, "unit": "a/b"},
    "BOSTWICK_cm":  {"min": 2.0,  "max": 6.0,  "target": 4.0,  "unit": "cm/30s"},
    "pH":           {"min": 4.10, "max": 4.45, "target": 4.20, "unit": ""},
    "ACIDITY_pct":  {"min": 0.0,  "max": 8.0,  "target": 6.5,  "unit": "%"},
    "BLOTTER_TEST": {"min": 0.0,  "max": 5.0,  "target": 3.0,  "unit": "mm"},
    "SYNERESIS_pct":{"min": 0.0,  "max": 6.0,  "target": 4.0,  "unit": "%"},
    "SALT_pct":     {"min": 1.5,  "max": 2.5,  "target": 2.0,  "unit": "%"},
}

# Κόστη — DEFENSIBLE για ελληνική μικρομεσαία γραμμή τομάτας (Kyknos-scale, 1 line)
# Παραδοχή: 3 t/h aseptic line, paste €1.150/tn χονδρική, 90-day Greek season
COST_CORRECTION   = 450     # €/παρτίδα — citric acid + labor + energy + QC + short downtime
COST_CRITICAL     = 6000    # €/παρτίδα — lost batch €3.500 + sanitization €800 + investigation €400 + disposal €300 + downstream €800 + NCR €200 (ratio 1.71× batch value)
BATCH_VALUE       = 3500    # €/παρτίδα — 3 τόνοι paste × €1.150/τόνο (wholesale EU 2024-25)

# HACCP / ISO 22000 reduction factors (real-world D. Nomikos QC experience)
HACCP_WASTE_REDUCTION = 0.25   # -25% σε raw material deviations & waste
HACCP_OPCOST_REDUCTION = 0.15  # -15% operational cost μέσω Brix/pH/Acidity/Colour QC
SEASON_MONTHS = 3              # Greek tomato season: 60-90 days/year (Aug-Oct), όχι 12 μήνες

# -----------------------------------------------------------------------------
# 2. DATA GENERATION — 24 μετρήσεις/ημέρα × 30 ημέρες = 720
# -----------------------------------------------------------------------------
def generate_data(start_date="2025-06-25", n_days=30, seed=42):
    """
    Default start: 25 Ιουνίου — early harvest για ελληνική τομάτα.
    Πλήρης σεζόνα: 25 Ιουν → 25 Σεπτ (≈ 90 ημέρες). Default = 30 days (1ος μήνας σεζόνας).
    """
    np.random.seed(seed)
    n = 24 * n_days
    start = datetime.fromisoformat(start_date)

    rows = []
    for h in range(n):
        ts = start + timedelta(hours=h)
        batch_id = f"B{h+1:04d}"

        # Realistic noise — οι περισσότερες μετρήσεις εντός spec,
        # ~10% λίγο εκτός (correction), ~3% κρίσιμα εκτός (waste)
        roll = np.random.random()

        if roll < 0.85:        # 85% Accept
            brix = np.random.normal(28.6, 0.35)
            color = np.random.normal(2.05, 0.08)
            bostwick = np.random.normal(4.0, 0.6)
            ph = np.random.normal(4.22, 0.06)
            acidity = np.random.normal(6.5, 0.5)
            blotter = np.random.normal(3.0, 0.6)
            syneresis = np.random.normal(3.8, 0.7)
            salt = np.random.normal(2.0, 0.15)
        elif roll < 0.97:      # 12% Correction (1-2 παράμετροι οριακά)
            brix = np.random.normal(28.6, 0.5)
            color = np.random.normal(1.95, 0.15)
            bostwick = np.random.normal(4.5, 1.4)        # μπορεί να ξεφύγει
            ph = np.random.normal(4.30, 0.18)            # συχνά οριακό
            acidity = np.random.normal(7.5, 0.8)
            blotter = np.random.normal(4.0, 1.0)
            syneresis = np.random.normal(5.0, 1.2)
            salt = np.random.normal(2.0, 0.25)
        else:                  # 3% Critical Waste
            brix = np.random.normal(26.5, 1.5)           # εκτός spec
            color = np.random.normal(1.70, 0.15)
            bostwick = np.random.normal(7.5, 1.5)
            ph = np.random.normal(4.55, 0.20)            # > 4.45 critical
            acidity = np.random.normal(9.5, 1.0)
            blotter = np.random.normal(6.5, 1.5)
            syneresis = np.random.normal(7.5, 1.5)
            salt = np.random.normal(2.3, 0.4)

        rows.append({
            "Date":          ts.strftime("%Y-%m-%d"),
            "Time":          ts.strftime("%H:%M"),
            "Day_of_week":   ts.strftime("%A"),
            "Timestamp":     ts.strftime("%Y-%m-%d %H:%M"),
            "Batch_ID":      batch_id,
            "BRIX":          round(max(0, brix), 2),
            "COLOR_a_b":     round(max(0, color), 2),
            "BOSTWICK_cm":   round(max(0, bostwick), 2),
            "pH":            round(max(0, ph), 2),
            "ACIDITY_pct":   round(max(0, acidity), 2),
            "BLOTTER_TEST":  round(max(0, blotter), 2),
            "SYNERESIS_pct": round(max(0, syneresis), 2),
            "SALT_pct":      round(max(0, salt), 2),
        })
    return pd.DataFrame(rows)


# -----------------------------------------------------------------------------
# 3. BATCH CLASSIFICATION & COSTING
# -----------------------------------------------------------------------------
def classify_row(row):
    """
    Επιστρέφει (status, cost, deviations_list)
    - Accept           : όλες οι παράμετροι εντός spec → 0€
    - Correction       : 1-2 παράμετροι οριακά εκτός → 450€
    - Critical Waste   : κρίσιμη παράμετρος (BRIX/pH) ή 3+ εκτός → 6000€
    """
    deviations = []
    critical_params = {"BRIX", "pH"}
    critical_hit = False

    for param, lim in SPEC.items():
        val = row[param]
        if val < lim["min"] or val > lim["max"]:
            deviations.append(param)
            # Έντονη παρέκκλιση σε κρίσιμη παράμετρο = critical waste
            tol = (lim["max"] - lim["min"]) * 0.15
            if param in critical_params and (
                val < lim["min"] - tol or val > lim["max"] + tol
            ):
                critical_hit = True

    if critical_hit or len(deviations) >= 3:
        return "Critical Waste", COST_CRITICAL, deviations
    elif len(deviations) >= 1:
        return "Correction", COST_CORRECTION, deviations
    else:
        return "Accept", 0, deviations


def process(df):
    statuses, costs, devs = [], [], []
    for _, row in df.iterrows():
        s, c, d = classify_row(row)
        statuses.append(s)
        costs.append(c)
        devs.append(", ".join(d) if d else "")
    df["Status"]            = statuses
    df["Action_Cost_EUR"]   = costs
    df["Deviations"]        = devs
    return df


# -----------------------------------------------------------------------------
# 4. REPORTING
# -----------------------------------------------------------------------------
def summarize(df):
    total = len(df)
    counts = df["Status"].value_counts().to_dict()
    accept    = counts.get("Accept", 0)
    correct   = counts.get("Correction", 0)
    waste     = counts.get("Critical Waste", 0)

    correction_cost = df.loc[df["Status"] == "Correction", "Action_Cost_EUR"].sum()
    waste_cost      = df.loc[df["Status"] == "Critical Waste", "Action_Cost_EUR"].sum()
    total_cost      = correction_cost + waste_cost

    revenue_lost    = waste * BATCH_VALUE
    yield_pct       = (accept + correct) / total * 100

    return {
        "Total batches":             total,
        "Accepted":                  accept,
        "Correction (€450)":         correct,
        "Critical Waste (€6000)":    waste,
        "Correction cost (€)":       correction_cost,
        "Waste cost (€)":            waste_cost,
        "Total financial impact (€)":total_cost,
        "Revenue lost (€)":          revenue_lost,
        "Production yield (%)":      round(yield_pct, 2),
    }


def haccp_simulation(df, summary):
    """
    Συγκρίνει Baseline (current data) vs Optimized (μετά εφαρμογή HACCP/ISO 22000).
    HACCP/ISO 22000 μειώνει waste/deviations -25% και operational cost -15%.
    """
    total = summary["Total batches"]
    waste = summary["Critical Waste (€6000)"]
    correct = summary["Correction (€450)"]
    accept = summary["Accepted"]

    # Optimized = -25% σε waste & corrections, οι παρτίδες αυτές πάνε σε Accept
    waste_after   = round(waste   * (1 - HACCP_WASTE_REDUCTION))
    correct_after = round(correct * (1 - HACCP_WASTE_REDUCTION))
    saved_batches = (waste - waste_after) + (correct - correct_after)
    accept_after  = accept + saved_batches

    correction_cost_before = correct * COST_CORRECTION
    waste_cost_before      = waste   * COST_CRITICAL
    total_cost_before      = correction_cost_before + waste_cost_before

    correction_cost_after  = correct_after * COST_CORRECTION * (1 - HACCP_OPCOST_REDUCTION)
    waste_cost_after       = waste_after   * COST_CRITICAL
    total_cost_after       = correction_cost_after + waste_cost_after

    monthly_savings = total_cost_before - total_cost_after
    annual_savings  = monthly_savings * SEASON_MONTHS

    revenue_recovered = (waste - waste_after) * BATCH_VALUE * SEASON_MONTHS

    yield_before = (accept + correct) / total * 100
    yield_after  = (accept_after + correct_after) / total * 100

    comparison = pd.DataFrame([
        {"Metric": "Total batches",              "Baseline": total,                "Optimized (HACCP)": total,                "Δ": 0},
        {"Metric": "Accepted",                   "Baseline": accept,               "Optimized (HACCP)": accept_after,         "Δ": accept_after - accept},
        {"Metric": "Correction (€450)",          "Baseline": correct,              "Optimized (HACCP)": correct_after,        "Δ": correct_after - correct},
        {"Metric": "Critical Waste (€6000)",     "Baseline": waste,                "Optimized (HACCP)": waste_after,          "Δ": waste_after - waste},
        {"Metric": "Production yield (%)",       "Baseline": round(yield_before,2),"Optimized (HACCP)": round(yield_after,2), "Δ": round(yield_after-yield_before,2)},
        {"Metric": "Correction cost (€)",        "Baseline": correction_cost_before,"Optimized (HACCP)": round(correction_cost_after,2),"Δ": round(correction_cost_after-correction_cost_before,2)},
        {"Metric": "Waste cost (€)",             "Baseline": waste_cost_before,    "Optimized (HACCP)": waste_cost_after,     "Δ": waste_cost_after - waste_cost_before},
        {"Metric": "TOTAL monthly cost (€)",     "Baseline": total_cost_before,    "Optimized (HACCP)": round(total_cost_after,2),"Δ": round(total_cost_after-total_cost_before,2)},
        {"Metric": "Monthly savings (€)",        "Baseline": 0,                    "Optimized (HACCP)": round(monthly_savings,2),"Δ": round(monthly_savings,2)},
        {"Metric": "Annual savings projection (€)","Baseline": 0,                  "Optimized (HACCP)": round(annual_savings,2),  "Δ": round(annual_savings,2)},
        {"Metric": "Annual revenue recovered (€)","Baseline": 0,                   "Optimized (HACCP)": round(revenue_recovered,2),"Δ": round(revenue_recovered,2)},
    ])

    # Risk reduction matrix ανά παράμετρο
    param_devs = {p: 0 for p in SPEC}
    for d in df["Deviations"]:
        if isinstance(d, str) and d.strip():
            for p in d.split(", "):
                if p in param_devs:
                    param_devs[p] += 1

    risk_matrix = pd.DataFrame([
        {
            "Parameter": p,
            "Baseline deviations": cnt,
            "Optimized deviations (-25%)": round(cnt * (1 - HACCP_WASTE_REDUCTION)),
            "Risk level": ("HIGH" if cnt > 50 else "MEDIUM" if cnt > 20 else "LOW"),
            "Control method": {
                "BRIX":          "Inline refractometer, citric acid dosing",
                "COLOR_a_b":     "Hunter Lab spectrophotometer, raw material screening",
                "BOSTWICK_cm":   "Viscosity meter, consistency control",
                "pH":            "Inline pH probe + citric acid auto-correction",
                "ACIDITY_pct":   "Titration QC every hour",
                "BLOTTER_TEST":  "Visual inspection per shift",
                "SYNERESIS_pct": "Stabilizer dosing, processing temp control",
                "SALT_pct":      "Conductivity meter, salt dosing PLC",
            }[p],
        }
        for p, cnt in param_devs.items()
    ])

    notes = pd.DataFrame([
        {"Note": "HACCP = Hazard Analysis Critical Control Points (μηδενική ανοχή σε κρίσιμα CCPs)"},
        {"Note": "ISO 22000 = Διεθνές πρότυπο Food Safety Management System"},
        {"Note": f"Παραδοχή: -{int(HACCP_WASTE_REDUCTION*100)}% deviations & waste μέσω CCP monitoring"},
        {"Note": f"Παραδοχή: -{int(HACCP_OPCOST_REDUCTION*100)}% operational cost μέσω physicochemical QC (Brix/pH/Acidity/Colour)"},
        {"Note": f"Annual projection = monthly savings × {SEASON_MONTHS}"},
        {"Note": "Πηγή ποσοστών: εμπειρία D. Nomikos QC role"},
    ])

    return comparison, risk_matrix, notes


def make_charts(df, summary, charts_dir):
    """Δημιουργεί όλα τα PNG charts στο φάκελο charts/."""
    os.makedirs(charts_dir, exist_ok=True)
    palette = {"Accept": "#2ecc71", "Correction": "#f39c12", "Critical Waste": "#e74c3c"}

    # 1. Status distribution (donut)
    fig, ax = plt.subplots(figsize=(7, 5))
    counts = df["Status"].value_counts().reindex(["Accept","Correction","Critical Waste"]).fillna(0)
    ax.pie(counts, labels=counts.index, colors=[palette[s] for s in counts.index],
           autopct="%1.1f%%", wedgeprops=dict(width=0.45), startangle=90, pctdistance=0.78)
    ax.set_title(f"Batch Status Distribution — {summary['Total batches']} hourly batches",
                 fontsize=13, fontweight="bold", pad=15)
    plt.tight_layout()
    plt.savefig(os.path.join(charts_dir, "01_status_distribution.png"), dpi=110, bbox_inches="tight")
    plt.close()

    # 2. Cost waterfall (correction vs waste vs total)
    fig, ax = plt.subplots(figsize=(8, 5))
    cats = ["Correction cost", "Waste cost", "TOTAL monthly"]
    vals = [summary["Correction cost (€)"], summary["Waste cost (€)"], summary["Total financial impact (€)"]]
    colors = ["#f39c12", "#e74c3c", "#34495e"]
    bars = ax.bar(cats, vals, color=colors, edgecolor="black")
    for b, v in zip(bars, vals):
        ax.text(b.get_x()+b.get_width()/2, b.get_height(), f"€{v:,.0f}",
                ha="center", va="bottom", fontweight="bold", fontsize=11)
    ax.set_title("Monthly Cost Breakdown (Baseline)", fontsize=13, fontweight="bold", pad=15)
    ax.set_ylabel("€")
    ax.grid(axis="y", alpha=0.3)
    ax.set_ylim(0, max(vals) * 1.15)
    plt.tight_layout()
    plt.savefig(os.path.join(charts_dir, "02_cost_breakdown.png"), dpi=110, bbox_inches="tight")
    plt.close()

    # 3. Daily trend (batches by status + daily cost)
    df["Date"] = pd.to_datetime(df["Timestamp"]).dt.date
    daily = df.groupby("Date").agg(
        Accept=("Status", lambda s: (s=="Accept").sum()),
        Correction=("Status", lambda s: (s=="Correction").sum()),
        Waste=("Status", lambda s: (s=="Critical Waste").sum()),
        Cost=("Action_Cost_EUR", "sum"),
    ).reset_index()
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(11, 7), sharex=True)
    ax1.bar(daily["Date"], daily["Accept"], color=palette["Accept"], label="Accept")
    ax1.bar(daily["Date"], daily["Correction"], bottom=daily["Accept"],
            color=palette["Correction"], label="Correction")
    ax1.bar(daily["Date"], daily["Waste"], bottom=daily["Accept"]+daily["Correction"],
            color=palette["Critical Waste"], label="Critical Waste")
    ax1.set_ylabel("Batches/day"); ax1.legend(loc="upper right", fontsize=9)
    ax1.set_title("Daily Batch Outcomes & Cost", fontsize=13, fontweight="bold", pad=10)
    ax1.grid(axis="y", alpha=0.3)

    ax2.plot(daily["Date"], daily["Cost"], color="#c0392b", marker="o", linewidth=2)
    ax2.fill_between(daily["Date"], daily["Cost"], color="#e74c3c", alpha=0.2)
    ax2.set_ylabel("Daily cost (€)"); ax2.grid(alpha=0.3)
    ax2.set_xlabel("Date")
    plt.setp(ax2.get_xticklabels(), rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig(os.path.join(charts_dir, "03_daily_trend.png"), dpi=110, bbox_inches="tight")
    plt.close()

    # 4. Parameter distribution vs spec limits (8 panels)
    fig, axes = plt.subplots(2, 4, figsize=(16, 8))
    for ax, (param, lim) in zip(axes.flat, SPEC.items()):
        vals = df[param]
        ax.hist(vals, bins=30, color="#3498db", edgecolor="black", alpha=0.7)
        ax.axvline(lim["min"], color="red", linestyle="--", linewidth=1.5, label=f"min {lim['min']}")
        ax.axvline(lim["max"], color="red", linestyle="--", linewidth=1.5, label=f"max {lim['max']}")
        ax.axvline(lim["target"], color="green", linestyle="-", linewidth=2, label=f"target {lim['target']}")
        ax.set_title(f"{param} ({lim['unit']})", fontsize=10, fontweight="bold")
        ax.legend(fontsize=7, loc="upper right")
        ax.grid(alpha=0.3)
    fig.suptitle("Parameter Distributions vs Spec Limits (MFD-024)",
                 fontsize=14, fontweight="bold", y=1.00)
    plt.tight_layout()
    plt.savefig(os.path.join(charts_dir, "04_parameter_distributions.png"), dpi=100, bbox_inches="tight")
    plt.close()

    # 5. HACCP Impact comparison
    accept = summary["Accepted"]; correct = summary["Correction (€450)"]
    waste = summary["Critical Waste (€6000)"]
    waste_after = round(waste * (1 - HACCP_WASTE_REDUCTION))
    correct_after = round(correct * (1 - HACCP_WASTE_REDUCTION))
    accept_after = summary["Total batches"] - waste_after - correct_after

    correction_cost_before = correct * COST_CORRECTION
    waste_cost_before = waste * COST_CRITICAL
    correction_cost_after = correct_after * COST_CORRECTION * (1 - HACCP_OPCOST_REDUCTION)
    waste_cost_after = waste_after * COST_CRITICAL

    fig, (a1, a2) = plt.subplots(1, 2, figsize=(13, 5.5))
    cats = ["Accept", "Correction", "Waste"]
    before = [accept, correct, waste]
    after = [accept_after, correct_after, waste_after]
    x = np.arange(len(cats)); w = 0.35
    a1.bar(x - w/2, before, w, label="Baseline", color="#95a5a6", edgecolor="black")
    a1.bar(x + w/2, after, w, label="With HACCP", color="#27ae60", edgecolor="black")
    a1.set_xticks(x); a1.set_xticklabels(cats)
    a1.set_ylabel("Batches"); a1.set_title("Batch Outcomes — Baseline vs HACCP", fontweight="bold")
    a1.legend(); a1.grid(axis="y", alpha=0.3)
    for i, (b, a) in enumerate(zip(before, after)):
        a1.text(i - w/2, b, str(b), ha="center", va="bottom", fontsize=9)
        a1.text(i + w/2, a, str(a), ha="center", va="bottom", fontsize=9)

    cost_cats = ["Correction €", "Waste €", "TOTAL €"]
    cost_before = [correction_cost_before, waste_cost_before, correction_cost_before+waste_cost_before]
    cost_after = [correction_cost_after, waste_cost_after, correction_cost_after+waste_cost_after]
    x2 = np.arange(len(cost_cats))
    a2.bar(x2 - w/2, cost_before, w, label="Baseline", color="#e74c3c", edgecolor="black")
    a2.bar(x2 + w/2, cost_after, w, label="With HACCP", color="#27ae60", edgecolor="black")
    a2.set_xticks(x2); a2.set_xticklabels(cost_cats)
    a2.set_ylabel("€/month"); a2.set_title("Cost Impact — Baseline vs HACCP", fontweight="bold")
    a2.legend(); a2.grid(axis="y", alpha=0.3)
    for i, (b, a) in enumerate(zip(cost_before, cost_after)):
        a2.text(i - w/2, b, f"€{b:,.0f}", ha="center", va="bottom", fontsize=8)
        a2.text(i + w/2, a, f"€{a:,.0f}", ha="center", va="bottom", fontsize=8)

    plt.tight_layout()
    plt.savefig(os.path.join(charts_dir, "05_haccp_impact.png"), dpi=110, bbox_inches="tight")
    plt.close()

    # 6. Pipeline flow diagram
    fig, ax = plt.subplots(figsize=(13, 4.5))
    ax.axis("off")
    boxes = [
        (0.02, "CSV\n720 hourly\nmeasurements", "#3498db"),
        (0.20, "Parse 8 params\n(BRIX, COLOR, ...)", "#9b59b6"),
        (0.38, "Compare to\nMFD-024 spec", "#1abc9c"),
        (0.56, "Classify batch\nAccept/Correct/Waste", "#f39c12"),
        (0.74, "Compute cost\n€/batch", "#e74c3c"),
        (0.92, "Excel report\n+ HACCP sim", "#27ae60"),
    ]
    for x, text, color in boxes:
        ax.add_patch(plt.Rectangle((x, 0.35), 0.14, 0.35, facecolor=color,
                                    edgecolor="black", linewidth=1.5, alpha=0.85))
        ax.text(x + 0.07, 0.525, text, ha="center", va="center",
                fontsize=9, fontweight="bold", color="white")
    for x, _, _ in boxes[:-1]:
        ax.annotate("", xy=(x + 0.17, 0.525), xytext=(x + 0.14, 0.525),
                    arrowprops=dict(arrowstyle="->", lw=2, color="black"))
    ax.set_xlim(0, 1.08); ax.set_ylim(0, 1)
    ax.set_title("Pipeline Flow — Industrial Tomato QC", fontsize=14, fontweight="bold", pad=10)
    plt.tight_layout()
    plt.savefig(os.path.join(charts_dir, "00_pipeline_flow.png"), dpi=110, bbox_inches="tight")
    plt.close()

    # =========================================================================
    # INSIGHT-DRIVEN CHARTS (Statistical Process Control + Root-Cause Analysis)
    # =========================================================================

    # 7. PARETO — ποια παράμετρος ευθύνεται για τις περισσότερες αποτυχίες (80/20)
    param_failures = {p: 0 for p in SPEC}
    for d in df["Deviations"]:
        if isinstance(d, str) and d.strip():
            for p in d.split(", "):
                if p in param_failures:
                    param_failures[p] += 1
    pareto = pd.Series(param_failures).sort_values(ascending=False)
    cumpct = (pareto.cumsum() / pareto.sum() * 100)

    fig, ax1 = plt.subplots(figsize=(10, 5.5))
    bars = ax1.bar(pareto.index, pareto.values, color="#e74c3c", edgecolor="black", alpha=0.85)
    ax1.set_ylabel("Failures (count)", color="#e74c3c", fontweight="bold")
    ax1.tick_params(axis="y", labelcolor="#e74c3c")
    for b, v in zip(bars, pareto.values):
        ax1.text(b.get_x()+b.get_width()/2, b.get_height(), str(int(v)),
                 ha="center", va="bottom", fontsize=9, fontweight="bold")
    ax2 = ax1.twinx()
    ax2.plot(pareto.index, cumpct.values, color="#2c3e50", marker="o", linewidth=2.2)
    ax2.axhline(80, color="green", linestyle="--", alpha=0.6, label="80% threshold")
    ax2.set_ylabel("Cumulative %", color="#2c3e50", fontweight="bold")
    ax2.set_ylim(0, 110); ax2.tick_params(axis="y", labelcolor="#2c3e50")
    ax2.legend(loc="lower right")
    plt.title("PARETO — Failure Mode Analysis (80/20 Rule)\nFocus HACCP investment here ↓",
              fontsize=13, fontweight="bold", pad=15)
    plt.setp(ax1.get_xticklabels(), rotation=30, ha="right")
    plt.tight_layout()
    plt.savefig(os.path.join(charts_dir, "06_pareto_failures.png"), dpi=110, bbox_inches="tight")
    plt.close()

    # 8. CORRELATION HEATMAP — ποιές παράμετροι αποτυγχάνουν μαζί
    param_cols = list(SPEC.keys())
    corr = df[param_cols].corr()
    fig, ax = plt.subplots(figsize=(9, 7))
    im = ax.imshow(corr.values, cmap="RdBu_r", vmin=-1, vmax=1, aspect="auto")
    ax.set_xticks(range(len(param_cols))); ax.set_yticks(range(len(param_cols)))
    ax.set_xticklabels(param_cols, rotation=45, ha="right")
    ax.set_yticklabels(param_cols)
    for i in range(len(param_cols)):
        for j in range(len(param_cols)):
            v = corr.values[i, j]
            ax.text(j, i, f"{v:.2f}", ha="center", va="center",
                    color="white" if abs(v) > 0.4 else "black", fontsize=9, fontweight="bold")
    plt.colorbar(im, ax=ax, fraction=0.04, pad=0.04, label="Pearson r")
    plt.title("CORRELATION MATRIX — Parameter Co-movement\n(spot which params drift together)",
              fontsize=13, fontweight="bold", pad=15)
    plt.tight_layout()
    plt.savefig(os.path.join(charts_dir, "07_correlation_heatmap.png"), dpi=110, bbox_inches="tight")
    plt.close()

    # 9. SPC CONTROL CHARTS — Statistical Process Control (3σ limits) for pH & BRIX
    fig, axes = plt.subplots(2, 1, figsize=(13, 7))
    for ax, param in zip(axes, ["pH", "BRIX"]):
        vals = df[param].values
        mu = vals.mean(); sigma = vals.std()
        ucl, lcl = mu + 3*sigma, mu - 3*sigma
        ax.plot(range(len(vals)), vals, color="#3498db", linewidth=0.7, alpha=0.7)
        ax.axhline(mu,  color="green", linestyle="-",  label=f"μ={mu:.2f}")
        ax.axhline(ucl, color="red",   linestyle="--", label=f"UCL (3σ)={ucl:.2f}")
        ax.axhline(lcl, color="red",   linestyle="--", label=f"LCL (3σ)={lcl:.2f}")
        ax.axhline(SPEC[param]["max"], color="orange", linestyle=":", label=f"Spec max={SPEC[param]['max']}")
        ax.axhline(SPEC[param]["min"], color="orange", linestyle=":", label=f"Spec min={SPEC[param]['min']}")
        # Highlight out-of-control points
        ooc = (vals > ucl) | (vals < lcl)
        ax.scatter(np.where(ooc)[0], vals[ooc], color="red", s=30, zorder=5, label=f"OOC ({ooc.sum()})")
        ax.set_title(f"SPC Control Chart — {param} (Western Electric Rules)", fontweight="bold")
        ax.set_xlabel("Batch index (hourly)"); ax.set_ylabel(param)
        ax.legend(loc="upper right", fontsize=8); ax.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(charts_dir, "08_spc_control_chart.png"), dpi=110, bbox_inches="tight")
    plt.close()

    # 10. PROCESS CAPABILITY (Cp / Cpk) — quantify how well process meets spec
    cap_rows = []
    for p, lim in SPEC.items():
        vals = df[p]
        sigma = vals.std()
        mu = vals.mean()
        if sigma == 0:
            cp = cpk = float("nan")
        else:
            cp  = (lim["max"] - lim["min"]) / (6 * sigma)
            cpu = (lim["max"] - mu) / (3 * sigma)
            cpl = (mu - lim["min"]) / (3 * sigma)
            cpk = min(cpu, cpl)
        cap_rows.append({"Param": p, "Cp": cp, "Cpk": cpk})
    cap = pd.DataFrame(cap_rows)

    fig, ax = plt.subplots(figsize=(11, 5.5))
    x = np.arange(len(cap)); w = 0.38
    ax.bar(x - w/2, cap["Cp"],  w, label="Cp (potential)", color="#3498db", edgecolor="black")
    ax.bar(x + w/2, cap["Cpk"], w, label="Cpk (actual)",   color="#e67e22", edgecolor="black")
    ax.axhline(1.33, color="green", linestyle="--", label="World-class (≥1.33)")
    ax.axhline(1.00, color="orange", linestyle="--", label="Acceptable (≥1.00)")
    ax.axhline(0.67, color="red", linestyle="--", label="Inadequate (<0.67)")
    ax.set_xticks(x); ax.set_xticklabels(cap["Param"], rotation=30, ha="right")
    ax.set_ylabel("Capability Index"); ax.legend(loc="upper right", fontsize=8)
    ax.set_title("PROCESS CAPABILITY (Cp / Cpk) — Industry Six Sigma Benchmark",
                 fontweight="bold", pad=12)
    ax.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(charts_dir, "09_process_capability.png"), dpi=110, bbox_inches="tight")
    plt.close()

    # 11. TIME-OF-DAY HEATMAP — εντοπισμός κρίσιμων shift hours
    df["Hour"] = pd.to_datetime(df["Timestamp"]).dt.hour
    df["DayOfMonth"] = pd.to_datetime(df["Timestamp"]).dt.day
    pivot = df.pivot_table(index="DayOfMonth", columns="Hour",
                            values="Action_Cost_EUR", aggfunc="sum", fill_value=0)
    fig, ax = plt.subplots(figsize=(14, 6))
    im = ax.imshow(pivot.values, aspect="auto", cmap="YlOrRd")
    ax.set_xticks(range(24)); ax.set_xticklabels(range(24))
    ax.set_yticks(range(len(pivot.index))); ax.set_yticklabels(pivot.index)
    ax.set_xlabel("Hour of day"); ax.set_ylabel("Day of month")
    plt.colorbar(im, ax=ax, label="QC failure cost (€)")
    plt.title("FAILURE HEATMAP — When do problems happen?\n(spot shift-pattern issues)",
              fontweight="bold", pad=12)
    plt.tight_layout()
    plt.savefig(os.path.join(charts_dir, "10_failure_heatmap.png"), dpi=110, bbox_inches="tight")
    plt.close()

    # 12. BOXPLOTS BY STATUS — distribution of each param by batch outcome
    fig, axes = plt.subplots(2, 4, figsize=(16, 8))
    order = ["Accept", "Correction", "Critical Waste"]
    for ax, param in zip(axes.flat, SPEC.keys()):
        data = [df[df["Status"] == s][param].values for s in order]
        bp = ax.boxplot(data, tick_labels=order, patch_artist=True, showfliers=True)
        for patch, color in zip(bp["boxes"], [palette[s] for s in order]):
            patch.set_facecolor(color); patch.set_alpha(0.7)
        ax.axhline(SPEC[param]["min"], color="red", linestyle="--", linewidth=1, alpha=0.7)
        ax.axhline(SPEC[param]["max"], color="red", linestyle="--", linewidth=1, alpha=0.7)
        ax.set_title(param, fontweight="bold")
        ax.tick_params(axis="x", rotation=20)
        ax.grid(axis="y", alpha=0.3)
    fig.suptitle("PARAMETER DISTRIBUTION BY OUTCOME (boxplots)\nWhich values predict batch rejection?",
                 fontsize=13, fontweight="bold", y=1.00)
    plt.tight_layout()
    plt.savefig(os.path.join(charts_dir, "11_boxplots_by_status.png"), dpi=100, bbox_inches="tight")
    plt.close()

    # 13. ROI / BREAK-EVEN — HACCP investment payback
    haccp_investment = 35000  # ένδειξη: HACCP implementation €5K-15K + ISO 22000 €15K-25K
    monthly_savings_val = summary["Correction cost (€)"]*HACCP_OPCOST_REDUCTION + \
                          (summary["Correction (€450)"]*COST_CORRECTION +
                           summary["Critical Waste (€6000)"]*COST_CRITICAL) * HACCP_WASTE_REDUCTION
    months = np.arange(0, 13)
    cum_savings = months * monthly_savings_val
    payback_month = haccp_investment / monthly_savings_val if monthly_savings_val > 0 else 0

    fig, ax = plt.subplots(figsize=(11, 5.5))
    ax.plot(months, cum_savings, color="#27ae60", linewidth=2.5, marker="o", label="Cumulative HACCP savings")
    ax.axhline(haccp_investment, color="red", linestyle="--", label=f"HACCP investment €{haccp_investment:,}")
    ax.axvline(payback_month, color="orange", linestyle=":", linewidth=2,
               label=f"Break-even ≈ {payback_month:.1f} months")
    ax.fill_between(months, cum_savings, haccp_investment,
                    where=(cum_savings >= haccp_investment), color="green", alpha=0.15, label="Net profit zone")
    ax.set_xlabel("Months since HACCP implementation"); ax.set_ylabel("€")
    ax.legend(loc="upper left"); ax.grid(alpha=0.3)
    ax.set_title("ROI & BREAK-EVEN — HACCP Investment Payback Analysis",
                 fontweight="bold", pad=12)
    plt.tight_layout()
    plt.savefig(os.path.join(charts_dir, "12_roi_breakeven.png"), dpi=110, bbox_inches="tight")
    plt.close()

    # 14. RISK PRIORITY MATRIX — Frequency × Severity
    severity = {"BRIX": 9, "pH": 10, "ACIDITY_pct": 8, "COLOR_a_b": 6,
                "BOSTWICK_cm": 5, "SYNERESIS_pct": 4, "BLOTTER_TEST": 3, "SALT_pct": 5}
    risk_data = []
    for p in SPEC:
        f = param_failures[p]
        s = severity[p]
        rpn = f * s
        risk_data.append((p, f, s, rpn))
    rdf = pd.DataFrame(risk_data, columns=["Param", "Frequency", "Severity", "RPN"])

    fig, ax = plt.subplots(figsize=(10, 7))
    sc = ax.scatter(rdf["Frequency"], rdf["Severity"],
                    s=rdf["RPN"]*4, c=rdf["RPN"], cmap="RdYlGn_r",
                    edgecolor="black", alpha=0.8)
    for _, row in rdf.iterrows():
        ax.annotate(row["Param"], (row["Frequency"], row["Severity"]),
                    xytext=(7, 7), textcoords="offset points", fontsize=9, fontweight="bold")
    # Quadrants
    fx = rdf["Frequency"].median(); sy = rdf["Severity"].median()
    ax.axvline(fx, color="gray", linestyle=":", alpha=0.5)
    ax.axhline(sy, color="gray", linestyle=":", alpha=0.5)
    ax.text(rdf["Frequency"].max()*0.95, rdf["Severity"].max()*0.98,
            "CRITICAL\n(monitor + control)", ha="right", va="top",
            fontweight="bold", color="darkred", fontsize=10,
            bbox=dict(boxstyle="round", fc="white", alpha=0.7))
    plt.colorbar(sc, ax=ax, label="Risk Priority Number (RPN)")
    ax.set_xlabel("Frequency (number of failures)")
    ax.set_ylabel("Severity (1-10 scale, food-safety impact)")
    ax.set_title("RISK PRIORITY MATRIX — Where to invest HACCP controls",
                 fontweight="bold", pad=12)
    ax.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(charts_dir, "13_risk_priority_matrix.png"), dpi=110, bbox_inches="tight")
    plt.close()

    return [
        "00_pipeline_flow.png",
        "01_status_distribution.png",
        "02_cost_breakdown.png",
        "03_daily_trend.png",
        "04_parameter_distributions.png",
        "05_haccp_impact.png",
        "06_pareto_failures.png",
        "07_correlation_heatmap.png",
        "08_spc_control_chart.png",
        "09_process_capability.png",
        "10_failure_heatmap.png",
        "11_boxplots_by_status.png",
        "12_roi_breakeven.png",
        "13_risk_priority_matrix.png",
    ]


def write_excel(df, summary, out_path, charts_dir=None):
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Hourly_Data", index=False)
        # Summary
        pd.DataFrame(list(summary.items()),
                     columns=["Metric", "Value"]).to_excel(
            w, sheet_name="Summary", index=False)
        # Spec limits sheet
        spec_df = pd.DataFrame([
            {"Parameter": k, **v} for k, v in SPEC.items()
        ])
        spec_df.to_excel(w, sheet_name="Spec_Limits", index=False)
        # Daily aggregation
        df["Date"] = pd.to_datetime(df["Timestamp"]).dt.date
        daily = df.groupby("Date").agg(
            Total_batches=("Batch_ID", "count"),
            Accepted=("Status", lambda s: (s == "Accept").sum()),
            Correction=("Status", lambda s: (s == "Correction").sum()),
            Waste=("Status", lambda s: (s == "Critical Waste").sum()),
            Cost_EUR=("Action_Cost_EUR", "sum"),
        ).reset_index()
        daily.to_excel(w, sheet_name="Daily_Summary", index=False)

        # HACCP Impact Simulation
        comparison, risk_matrix, notes = haccp_simulation(df, summary)
        comparison.to_excel(w, sheet_name="HACCP_Impact", index=False, startrow=2)
        risk_matrix.to_excel(w, sheet_name="HACCP_Impact", index=False,
                              startrow=len(comparison) + 6)
        notes.to_excel(w, sheet_name="HACCP_Impact", index=False,
                        startrow=len(comparison) + len(risk_matrix) + 10)

        # Title rows
        ws = w.sheets["HACCP_Impact"]
        ws["A1"] = "HACCP / ISO 22000 IMPACT SIMULATION — Baseline vs Optimized"
        ws["A1"].font = ws["A1"].font.copy(bold=True, size=14)
        ws.cell(row=len(comparison) + 6, column=1).value = "RISK REDUCTION MATRIX (per parameter)"
        ws.cell(row=len(comparison) + 6, column=1).font = ws.cell(row=len(comparison) + 6, column=1).font.copy(bold=True, size=12)
        ws.cell(row=len(comparison) + len(risk_matrix) + 10, column=1).value = "NOTES & ASSUMPTIONS"
        ws.cell(row=len(comparison) + len(risk_matrix) + 10, column=1).font = ws.cell(row=len(comparison) + len(risk_matrix) + 10, column=1).font.copy(bold=True, size=12)
        # Column widths
        for col, w_ in zip("ABCD", [42, 18, 22, 18]):
            ws.column_dimensions[col].width = w_

        # Embed charts as a new sheet
        if charts_dir and os.path.isdir(charts_dir):
            wb = w.book
            ws_charts = wb.create_sheet("Charts")
            ws_charts["A1"] = "VISUAL DASHBOARD — Industrial Tomato QC"
            ws_charts["A1"].font = ws_charts["A1"].font.copy(bold=True, size=16)
            row = 3
            for fname in sorted(os.listdir(charts_dir)):
                if fname.endswith(".png"):
                    img_path = os.path.join(charts_dir, fname)
                    img = XLImage(img_path)
                    img.width  = int(img.width  * 0.55)
                    img.height = int(img.height * 0.55)
                    ws_charts.add_image(img, f"A{row}")
                    row += 28


# -----------------------------------------------------------------------------
# 5. MAIN
# -----------------------------------------------------------------------------
if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    csv_path   = os.path.join(here, "tomato_qc_data.csv")
    xlsx_path  = os.path.join(here, "tomato_qc_report.xlsx")

    # Αν υπάρχει ήδη CSV → φόρτωσέ το (ώστε να μπορείς να φορτώσεις δικά σου data)
    # Force regenerate: delete tomato_qc_data.csv πριν τρέξεις ή χρησιμοποίησε --regen
    import sys
    force_regen = "--regen" in sys.argv

    if os.path.exists(csv_path) and not force_regen:
        print(f"Loading existing CSV: {csv_path}")
        df = pd.read_csv(csv_path)
    else:
        print("Generating synthetic hourly data (24/day x 30 days = 720)...")
        print("Start date: 25 Ιουνίου 2025 (early Greek tomato harvest)")
        df = generate_data()
        df.to_csv(csv_path, index=False, encoding="utf-8")
        print(f"Saved: {csv_path}")

    df = process(df)
    summary = summarize(df)

    print("\n" + "="*52)
    print("   PRODUCTION QUALITY FINANCIAL REPORT")
    print("="*52)
    for k, v in summary.items():
        print(f"  {k:30s} {v}")
    print("="*52)

    # Επανεγγραφή CSV με τα status/cost columns
    df.to_csv(csv_path, index=False, encoding="utf-8")

    charts_dir = os.path.join(here, "charts")
    print("\nGenerating charts...")
    chart_files = make_charts(df, summary, charts_dir)
    for c in chart_files:
        print(f"  - charts/{c}")

    write_excel(df, summary, xlsx_path, charts_dir=charts_dir)
    print(f"\nCSV  : {csv_path}")
    print(f"Excel: {xlsx_path}")
    print(f"Charts dir: {charts_dir}")
